from openpyxl import Workbook
from openpyxl import load_workbook
from django.utils import timezone
from tablib import Dataset
from .models import ATTENDANCE_INFO

# Create a new Excel workbook
# wb = Workbook()

# Select the active worksheet
# ws = wb.active

current_date = timezone.now().date()
day_week = current_date.weekday()
print(day_week)
days={
     0 : "MONDAY",
     1 : "TUESDAY",
     2 : "WEDNESDAY",
     3 : "THURSDAY",
     4 : "FRIDAY"
}

dataset = Dataset()
imported_data = dataset.load(open('excels/timetable.xlsx', 'rb').read(),format='xlsx')

subjects =[]
for data in imported_data:
    if data[0] == days[day_week]:
        for i in range(1,9):
            subjects.append(data[i])
print(subjects)
wb = load_workbook('excels/attendance.xlsx')

# Iterate over all sheets in the workbook
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # Delete all rows in the sheet
    ws.delete_rows(1, ws.max_row)

# Define the headers
wb.append(["ATTENDANCE_ID","DATE","HOUR","SUBJECT_ID","STATUS"])

all_objects = ATTENDANCE_INFO.objects.all()

# Iterate over the objects and access their fields
for obj in all_objects:
    for i in range(1,9):
        wb.append(obj.ATTENDANCE_ID,current_date,i,subjects[i-1],"PRESENT") 

wb.save('excels/attendance.xlsx')
print("created")
# with open('excels/attendance.xlsx', 'rb') as excel_file:
#         response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
#         return response


# Save the workbook to a file
